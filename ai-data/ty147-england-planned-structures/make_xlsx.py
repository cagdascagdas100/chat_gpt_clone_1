import csv, sys
from pathlib import Path
csv_path = Path(r'C:\AAYS1_GITHUB_BRIDGE\chat_gpt_clone_1\ai-results\ty147-england-planned-structures-long-harvest.master.csv')
xlsx_path = Path(r'C:\AAYS1_GITHUB_BRIDGE\chat_gpt_clone_1\ai-results\ty147-england-planned-structures-long-harvest.xlsx')
try:
    import openpyxl
except Exception as e:
    print('openpyxl_missing')
    sys.exit(2)
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Projeler'
with csv_path.open('r', encoding='utf-8-sig', newline='') as f:
    for row in csv.reader(f):
        ws.append(row)
for cell in ws[1]:
    cell.font = openpyxl.styles.Font(bold=True)
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
wb.create_sheet('Skalalar')
sh = wb['Skalalar']
sh.append(['Skala','Anlam'])
sh.append([1,'DÃ¼ÅŸÃ¼k: belirsiz veya yalnÄ±zca liste/erken aÅŸama'])
sh.append([2,'Orta: proje/region dÃ¼zeyi; boundary gerekir'])
sh.append([3,'YÃ¼ksek: resmi proje sayfasÄ± veya named corridor'])
sh.append([4,'Ã‡ok yÃ¼ksek: resmi karar/boundary/koordinat verisi'])
wb.save(xlsx_path)
print('xlsx_created')
