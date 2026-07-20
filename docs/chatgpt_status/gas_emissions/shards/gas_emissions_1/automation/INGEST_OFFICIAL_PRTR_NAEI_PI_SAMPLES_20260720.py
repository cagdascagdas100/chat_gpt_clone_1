from __future__ import annotations

import concurrent.futures
import csv
import hashlib
import html
import io
import json
import os
import shutil
import tempfile
import urllib.request
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from xml.etree import ElementTree as ET

SLOT_ID = "gas_emissions_1"
SOURCES = [
    {
        "source_id": "UK_PRTR_2024_XML",
        "url": "https://assets.publishing.service.gov.uk/media/6a3d096c4c7605ab56723a63/uk_prtr_dataset_2024.xml",
        "format": "xml",
        "publisher": "Defra",
        "published_date": "2026-06-30",
        "max_bytes": 40_000_000,
    },
    {
        "source_id": "NAEI_POINT_SOURCES_2023_XLSX",
        "url": "https://naei.energysecurity.gov.uk/sites/default/files/2025-09/NAEIPointsSources_2023.xlsx",
        "format": "xlsx",
        "publisher": "National Atmospheric Emissions Inventory",
        "published_date": "2025-09-30",
        "max_bytes": 140_000_000,
    },
    {
        "source_id": "EA_POLLUTION_INVENTORY_2024_ZIP",
        "url": "https://environment.data.gov.uk/api/file/download?fileDataSetId=4faa4a52-7df2-4047-bc3f-877dd04222d8&fileName=2024+Pollution+Inventory+Dataset.zip",
        "format": "zip",
        "publisher": "Environment Agency",
        "published_date": "2026-04-17",
        "max_bytes": 180_000_000,
    },
]


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def download(source: dict, target: Path) -> dict:
    request = urllib.request.Request(
        source["url"],
        headers={"User-Agent": "AAYS-TerraYield-official-source-ingestion/1.0"},
    )
    written = 0
    with urllib.request.urlopen(request, timeout=120) as response, target.open("wb") as output:
        while True:
            chunk = response.read(1024 * 1024)
            if not chunk:
                break
            written += len(chunk)
            if written > int(source["max_bytes"]):
                raise RuntimeError(f"SOURCE_SIZE_LIMIT_EXCEEDED:{source['source_id']}:{written}")
            output.write(chunk)
    return {
        "downloaded": True,
        "bytes": written,
        "sha256": sha256_file(target),
        "content_type": None,
    }


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1].casefold()


def parse_xml(path: Path, source_id: str, limit: int = 8) -> tuple[list[dict], dict]:
    samples: list[dict] = []
    tag_counts: dict[str, int] = {}
    facility_names = {"facilityname", "facility_name", "nameoffacility"}
    identifiers = {"facilityid", "nationalid", "facilityidentifier", "facility_id"}
    lat_names = {"latitude", "lat"}
    lon_names = {"longitude", "lon", "long"}
    record_tags = {"facilityreport", "facility", "productionfacility", "reportdata"}
    for _event, element in ET.iterparse(path, events=("end",)):
        name = local_name(element.tag)
        tag_counts[name] = tag_counts.get(name, 0) + 1
        if name in record_tags and len(samples) < limit:
            record = {"source_id": source_id, "facility_name": None, "facility_id": None, "latitude": None, "longitude": None}
            for child in element.iter():
                child_name = local_name(child.tag)
                value = (child.text or "").strip()
                if not value:
                    continue
                if child_name in facility_names and not record["facility_name"]:
                    record["facility_name"] = value
                elif child_name in identifiers and not record["facility_id"]:
                    record["facility_id"] = value
                elif child_name in lat_names and not record["latitude"]:
                    record["latitude"] = value
                elif child_name in lon_names and not record["longitude"]:
                    record["longitude"] = value
            if record["facility_name"] or record["facility_id"]:
                record.update({"data_status": "official_facility_candidate", "parcel_binding_status": "NOT_RUN"})
                samples.append(record)
        element.clear()
    common_tags = sorted(tag_counts.items(), key=lambda item: (-item[1], item[0]))[:30]
    return samples, {"top_tag_counts": common_tags, "sample_count": len(samples)}


def normalise_header(value: object) -> str:
    return "".join(ch for ch in str(value or "").casefold() if ch.isalnum())


def choose_columns(headers: list[object]) -> dict[str, int | None]:
    normalised = [normalise_header(value) for value in headers]
    def first(tokens: tuple[str, ...]) -> int | None:
        for index, header in enumerate(normalised):
            if any(token in header for token in tokens):
                return index
        return None
    return {
        "facility_name": first(("facilityname", "sitename", "installationname", "operatorname", "name")),
        "facility_id": first(("facilityid", "permitnumber", "registrationnumber", "nationalid", "siteid")),
        "easting": first(("easting", "eastings", "xcoordinate", "gridx")),
        "northing": first(("northing", "northings", "ycoordinate", "gridy")),
        "latitude": first(("latitude", "lat")),
        "longitude": first(("longitude", "lon", "long")),
        "pollutant": first(("pollutant", "substance")),
        "quantity": first(("quantity", "emission", "release", "amount")),
    }


def row_to_sample(row: list[object], columns: dict[str, int | None], source_id: str) -> dict | None:
    sample: dict[str, object] = {"source_id": source_id}
    for field, index in columns.items():
        sample[field] = None if index is None or index >= len(row) else row[index]
    if not sample.get("facility_name") and not sample.get("facility_id"):
        return None
    if not any(sample.get(field) not in (None, "") for field in ("easting", "northing", "latitude", "longitude")):
        return None
    sample.update({"data_status": "official_facility_candidate", "parcel_binding_status": "NOT_RUN"})
    return sample


def parse_xlsx(path: Path, source_id: str, limit: int = 8) -> tuple[list[dict], dict]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        return [], {"error": f"OPENPYXL_UNAVAILABLE:{exc}"}
    workbook = load_workbook(path, read_only=True, data_only=True)
    samples: list[dict] = []
    sheet_summaries: list[dict] = []
    for worksheet in workbook.worksheets:
        iterator = worksheet.iter_rows(values_only=True)
        header = None
        columns = None
        scanned = 0
        for row in iterator:
            scanned += 1
            if scanned > 30:
                break
            candidate = choose_columns(list(row))
            if candidate["facility_name"] is not None and any(candidate[key] is not None for key in ("easting", "northing", "latitude", "longitude")):
                header = list(row)
                columns = candidate
                break
        found = 0
        if header is not None and columns is not None:
            for row in iterator:
                sample = row_to_sample(list(row), columns, source_id)
                if sample:
                    sample["sheet"] = worksheet.title
                    samples.append(sample)
                    found += 1
                    if len(samples) >= limit:
                        break
        sheet_summaries.append({"sheet": worksheet.title, "header_found": header is not None, "sample_rows": found})
        if len(samples) >= limit:
            break
    workbook.close()
    return samples, {"sheets": sheet_summaries, "sample_count": len(samples)}


def parse_csv_bytes(data: bytes, source_id: str, limit: int = 8) -> tuple[list[dict], dict]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = iter(reader)
    headers = next(rows, [])
    columns = choose_columns(headers)
    samples = []
    for row in rows:
        sample = row_to_sample(row, columns, source_id)
        if sample:
            samples.append(sample)
            if len(samples) >= limit:
                break
    return samples, {"headers": headers[:50], "sample_count": len(samples)}


def parse_zip(path: Path, source_id: str, temp_root: Path, limit: int = 8) -> tuple[list[dict], dict]:
    samples: list[dict] = []
    member_summaries: list[dict] = []
    with zipfile.ZipFile(path) as archive:
        members = [name for name in archive.namelist() if not name.endswith("/")]
        for name in members:
            lower = name.casefold()
            if len(samples) >= limit:
                break
            try:
                if lower.endswith(".csv"):
                    data = archive.read(name)
                    part, meta = parse_csv_bytes(data, source_id, limit - len(samples))
                elif lower.endswith(".xlsx"):
                    extracted = temp_root / Path(name).name
                    with archive.open(name) as source, extracted.open("wb") as target:
                        shutil.copyfileobj(source, target)
                    part, meta = parse_xlsx(extracted, source_id, limit - len(samples))
                else:
                    continue
                for sample in part:
                    sample["archive_member"] = name
                samples.extend(part)
                member_summaries.append({"member": name, **meta})
            except Exception as exc:
                member_summaries.append({"member": name, "error": f"{type(exc).__name__}:{exc}"})
    return samples, {"members": member_summaries, "sample_count": len(samples)}


def process_source(source: dict, root: Path) -> dict:
    target = root / f"{source['source_id']}.{source['format']}"
    result = {**source, "download": None, "parse": None, "samples": [], "state": "RUNNING"}
    try:
        result["download"] = download(source, target)
        if source["format"] == "xml":
            samples, meta = parse_xml(target, source["source_id"])
        elif source["format"] == "xlsx":
            samples, meta = parse_xlsx(target, source["source_id"])
        elif source["format"] == "zip":
            samples, meta = parse_zip(target, source["source_id"], root)
        else:
            samples, meta = [], {"error": "UNSUPPORTED_FORMAT"}
        result["samples"] = samples
        result["parse"] = meta
        result["state"] = "PASS" if samples else "BLOCKED_NO_COORDINATED_SAMPLE"
    except Exception as exc:
        result["state"] = "BLOCKED"
        result["error"] = f"{type(exc).__name__}:{exc}"
    finally:
        try:
            target.unlink(missing_ok=True)
        except Exception:
            pass
    return result


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> int:
    slot_id = os.environ.get("AAYS_SLOT_ID", "")
    task_id = os.environ.get("AAYS_TASK_ID", "")
    if slot_id != SLOT_ID or not task_id:
        raise RuntimeError("GAS_EMISSIONS_1_OFFICIAL_INGEST_WRONG_SLOT_CONTEXT")
    root = Path.cwd()
    with tempfile.TemporaryDirectory(prefix="aays_gas_official_") as temporary:
        temp_root = Path(temporary)
        with concurrent.futures.ThreadPoolExecutor(max_workers=3) as pool:
            results = list(pool.map(lambda source: process_source(source, temp_root), SOURCES))
    samples = []
    for result in results:
        samples.extend(result.get("samples") or [])
    samples = samples[:18]
    pass_count = sum(result.get("state") == "PASS" for result in results)
    payload = {
        "schema_version": 1,
        "slot_id": SLOT_ID,
        "task_id": task_id,
        "parcel_partition": {"start": 1, "end": 30761, "count": 30761},
        "status": "PASS_OFFICIAL_SAMPLE_EXTRACTION" if pass_count >= 1 and samples else "BLOCKED_NO_OFFICIAL_COORDINATED_SAMPLES",
        "generated_at": utc_now(),
        "source_count": len(SOURCES),
        "source_pass_count": pass_count,
        "source_results": results,
        "sample_candidate_count": len(samples),
        "samples": samples,
        "measured_parcel_rows_created": 0,
        "parcel_values_created": 0,
        "parcel_binding_gate_passed": False,
        "next_action": "Verify coordinate CRS and join only exact official points or grid cells to verified parcel polygons.",
        "final_ready": False,
        "product_final_ready": False,
        "fake_data": False,
        "db_write": False,
        "migration": False,
        "production_deploy": False,
    }
    report = root / "docs/chatgpt_status/gas_emissions/shards/gas_emissions_1/reports/gas_emissions_1_official_ingestion_latest.json"
    status = root / "docs/chatgpt_status/gas_emissions/shards/gas_emissions_1/status/gas_emissions_1_official_ingestion_latest.json"
    web_json = root / "england_map_web/data/aays_21_slots/gas_emissions_1/official_facility_samples_latest.json"
    for path in (report, status, web_json):
        write_json(path, payload)
    rows = []
    for index, sample in enumerate(samples, start=1):
        rows.append(
            "<tr>"
            f"<td>{index}</td><td>{html.escape(str(sample.get('source_id') or ''))}</td>"
            f"<td>{html.escape(str(sample.get('facility_id') or ''))}</td>"
            f"<td>{html.escape(str(sample.get('facility_name') or ''))}</td>"
            f"<td>{html.escape(str(sample.get('easting') or sample.get('longitude') or ''))}</td>"
            f"<td>{html.escape(str(sample.get('northing') or sample.get('latitude') or ''))}</td>"
            f"<td>{html.escape(str(sample.get('pollutant') or ''))}</td>"
            f"<td>{html.escape(str(sample.get('quantity') or ''))}</td>"
            "<td>ADAY — parsel bağı çalıştırılmadı</td></tr>"
        )
    html_payload = f"""<!doctype html><html lang='tr'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>Gas Emissions 1 — Resmî tesis örnekleri</title><style>body{{font-family:Arial,sans-serif;margin:20px;background:#f5f7fa;color:#17202a}}table{{border-collapse:collapse;width:100%;background:white;font-size:13px}}th,td{{border:1px solid #cfd8dc;padding:7px;text-align:left;vertical-align:top}}th{{background:#eceff1;position:sticky;top:0}}.notice{{background:#fff3cd;border:1px solid #ffe69c;padding:12px;margin:12px 0}}</style></head><body><h1>gas_emissions_1 — Resmî tesis örnek adayları</h1><div class='notice'>Bu satırlar resmî kaynaklardan çıkarılan tesis adaylarıdır. Parsel emisyon ölçümü değildir; CRS ve gerçek parsel poligonu doğrulanmadan bağlama yapılmaz.</div><p>Kaynak: {len(SOURCES)} · Başarılı kaynak: {pass_count} · Aday satır: {len(samples)} · Ölçülmüş parsel satırı: 0</p><table><thead><tr><th>#</th><th>Kaynak</th><th>Tesis ID</th><th>Tesis</th><th>X/Lon</th><th>Y/Lat</th><th>Kirletici</th><th>Miktar</th><th>Durum</th></tr></thead><tbody>{''.join(rows)}</tbody></table></body></html>"""
    web_html = root / "england_map_web/data/aays_21_slots/gas_emissions_1/official_facility_samples.html"
    web_html.parent.mkdir(parents=True, exist_ok=True)
    web_html.write_text(html_payload, encoding="utf-8")
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0 if samples else 2


if __name__ == "__main__":
    raise SystemExit(main())
